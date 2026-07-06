import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime, date, timedelta
import streamlit as st


CSV_OS_ALLOWED = ["ibm z/os", "ibm z/vm"]

REPORT_COLUMNS = [
    "APAR/TPF/FMID (IBM)",
    "Severity (IBM)",
    "CVSS (IBM)",
    "Due Date (AXP)",
    "Escalation (AXP)",
    "sec_uuid",
    "cve_id",
    "host_name",
    "operating_system",
    "cvss",
    "cve_cache.published",
    "remediation_status",
    "remediation_comment",
    "Team",
]

APEXA_COLUMNS = [
    "cve_id",
    "remediation_status",
    "host_name",
    "device_uuid",
    "msp_tenant_uid",
    "remediation_comment",
]


def normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_date(value):
    if pd.isna(value):
        return ""
    return str(value).split("T")[0].strip()


def normalize_severity(value):
    value = normalize_text(value).upper()
    return "MEDIUM" if value == "MODERATE" else value


def normalize_remediation_status(value):
    value = normalize_text(value).upper()

    if value in ["", "DEFFERED", "DEFERRED"]:
        return "PENDING"

    return value


def normalize_remediation_comment(value):
    value = normalize_text(value)

    if value == "":
        return "Pending Status"

    return value


def parse_date(value):
    value = normalize_date(value)

    if not value:
        return None

    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def calculate_due_date(published, severity):
    published_date = parse_date(published)
    severity = normalize_severity(severity)

    if not published_date:
        return ""

    if severity == "CRITICAL":
        return published_date + timedelta(days=14)

    if severity == "HIGH":
        return published_date + timedelta(days=60)

    if severity == "MEDIUM":
        return published_date + timedelta(days=180)

    if severity == "LOW":
        return ""

    return ""


def is_overdue(published, severity):
    due_date = calculate_due_date(published, severity)

    if not due_date:
        return False

    return date.today() > due_date


def get_team(affected_products):
    text = normalize_text(affected_products).lower()

    has_mvs = (
        "cpe:2.3:o:ibm:z\\/os:" in text
        or "cpe:2.3:o:ibm:z/os:" in text
        or "z\\/os_security" in text
        or "z/os_security" in text
        or "z\\/os_system_automation" in text
        or "z/os_system_automation" in text
        or "z\\/os_storage_software" in text
        or "z/os_storage_software" in text
    )

    has_omegamon = "z\\/os_performance" in text or "z/os_performance" in text
    has_mq = "z\\/os_websphere_mq" in text or "z/os_websphere_mq" in text
    has_cics = "z\\/os_cics" in text or "z/os_cics" in text
    has_db2 = "z\\/os_db2" in text or "z/os_db2" in text
    has_ims = "z\\/os_ims" in text or "z/os_ims" in text
    has_zvm = "cpe:2.3:a:ibm:z\\/vm" in text or "cpe:2.3:a:ibm:z/vm" in text

    if has_db2 and has_ims:
        return "zOS DB2 & IMS"

    if has_mvs and has_omegamon:
        return "zOS MVS & zOS Omegamon"

    if has_mq:
        return "zOS MQ"

    if has_cics:
        return "zOS CICS"

    if has_db2:
        return "zOS DB2"

    if has_ims:
        return "zOS IMS"

    if has_zvm:
        return "zVM"

    if has_mvs:
        return "zOS MVS"

    if has_omegamon:
        return "zOS Omegamon"

    return "UNKNOWN"


def validate_columns(df, required_columns, file_name):
    missing = [col for col in required_columns if col not in df.columns]

    if missing:
        raise ValueError(f"Missing columns in {file_name}: {', '.join(missing)}")


def build_output_row(row):
    cve_id = normalize_text(row.get("cve_id"))
    host_name = normalize_text(row.get("host_name"))
    severity = normalize_severity(row.get("severity_name"))
    published = normalize_date(row.get("cve_cache.published"))
    due_date = calculate_due_date(published, severity)

    return {
        "APAR/TPF/FMID (IBM)": "",
        "Severity (IBM)": "",
        "CVSS (IBM)": "",
        "Due Date (AXP)": due_date.strftime("%Y-%m-%d") if due_date else "",
        "Escalation (AXP)": "",
        "sec_uuid": f"{cve_id}_{host_name}",
        "cve_id": cve_id,
        "host_name": host_name,
        "operating_system": normalize_text(row.get("operating_system")),
        "cvss": severity,
        "cve_cache.published": published,
        "remediation_status": normalize_remediation_status(row.get("remediation_status")),
        "remediation_comment": normalize_remediation_comment(row.get("remediation_comment")),
        "Team": get_team(row.get("affected_products")),
    }


def generate_report(report_file, csv_file):
    report_df = pd.read_excel(report_file, sheet_name=0, dtype=str)
    csv_df = pd.read_csv(csv_file, dtype=str)

    report_df.columns = report_df.columns.str.strip()
    csv_df.columns = csv_df.columns.str.strip()

    validate_columns(
        report_df,
        ["cve_id", "remediation_status", "remediation_comment"],
        "current report XLSX",
    )

    validate_columns(
        csv_df,
        [
            "cve_id",
            "host_name",
            "operating_system",
            "severity_name",
            "cve_cache.published",
            "remediation_status",
            "remediation_comment",
            "affected_products",
        ],
        "CSV",
    )

    csv_df["operating_system_norm"] = (
        csv_df["operating_system"]
        .fillna("")
        .str.strip()
        .str.lower()
    )

    csv_df = csv_df[csv_df["operating_system_norm"].isin(CSV_OS_ALLOWED)].copy()

    report_df["cve_id_norm"] = (
        report_df["cve_id"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    csv_df["cve_id_norm"] = (
        csv_df["cve_id"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    csv_df["remediation_status_norm"] = csv_df["remediation_status"].apply(
        normalize_remediation_status
    )

    csv_df["remediation_comment_norm"] = csv_df["remediation_comment"].apply(
        normalize_remediation_comment
    )

    report_df["remediation_status_norm"] = report_df["remediation_status"].apply(
        normalize_remediation_status
    )

    report_df["remediation_comment_norm"] = report_df["remediation_comment"].apply(
        normalize_remediation_comment
    )

    existing_cves = set(report_df["cve_id_norm"].dropna())

    new_releases_df = csv_df[
        ~csv_df["cve_id_norm"].isin(existing_cves)
        & (csv_df["cve_id_norm"] != "")
    ].copy()

    report_lookup = report_df.set_index("cve_id_norm")

    pending_rows = []
    apexa_rows = []

    for _, row in csv_df.iterrows():
        cve_id = row["cve_id_norm"]

        if cve_id in report_lookup.index:
            report_row = report_lookup.loc[cve_id]

            if isinstance(report_row, pd.DataFrame):
                report_row = report_row.iloc[0]

            csv_status = normalize_remediation_status(row.get("remediation_status"))
            csv_comment = normalize_remediation_comment(row.get("remediation_comment"))

            report_status = normalize_remediation_status(
                report_row.get("remediation_status")
            )
            report_comment = normalize_remediation_comment(
                report_row.get("remediation_comment")
            )

            if csv_status != report_status or csv_comment != report_comment:
                pending_rows.append({
                    "cve_id": cve_id,
                    "host_name": normalize_text(row.get("host_name")),
                    "operating_system": normalize_text(row.get("operating_system")),
                    "csv_remediation_status": csv_status,
                    "report_remediation_status": report_status,
                    "csv_remediation_comment": csv_comment,
                    "report_remediation_comment": report_comment,
                })

                apexa_rows.append({
                    "cve_id": cve_id,
                    "remediation_status": report_status,
                    "host_name": normalize_text(row.get("host_name")),
                    "device_uuid": "",
                    "msp_tenant_uid": "",
                    "remediation_comment": report_comment,
                })

    pending_updates_df = pd.DataFrame(pending_rows)

    apexa_update_df = pd.DataFrame(apexa_rows, columns=APEXA_COLUMNS)

    overdue_df = csv_df[
        (csv_df["remediation_status_norm"] == "PENDING")
        & csv_df.apply(
            lambda row: is_overdue(
                row.get("cve_cache.published"),
                row.get("severity_name"),
            ),
            axis=1,
        )
    ].copy()

    overdue_rows = []

    for _, row in overdue_df.iterrows():
        severity = normalize_severity(row.get("severity_name"))
        published = normalize_date(row.get("cve_cache.published"))
        due_date = calculate_due_date(published, severity)

        overdue_rows.append({
            "cve_id": normalize_text(row.get("cve_id")),
            "host_name": normalize_text(row.get("host_name")),
            "operating_system": normalize_text(row.get("operating_system")),
            "severity": severity,
            "published": published,
            "due_date": due_date.strftime("%Y-%m-%d") if due_date else "",
            "days_overdue": (date.today() - due_date).days if due_date else "",
            "remediation_status": normalize_remediation_status(row.get("remediation_status")),
            "remediation_comment": normalize_remediation_comment(row.get("remediation_comment")),
            "Team": get_team(row.get("affected_products")),
        })

    overdue_result_df = pd.DataFrame(overdue_rows)

    append_rows = [build_output_row(row) for _, row in new_releases_df.iterrows()]
    append_df = pd.DataFrame(append_rows, columns=REPORT_COLUMNS)

    total_open = csv_df[
        csv_df["remediation_status_norm"] == "PENDING"
    ]["cve_id_norm"].nunique()

    total_closed = csv_df[
        csv_df["remediation_status_norm"].isin(["REMEDIATED", "FALSE_POSITIVE"])
    ]["cve_id_norm"].nunique()

    pending_to_apply = total_open

    report_file.seek(0)
    wb = load_workbook(report_file)
    ws = wb.worksheets[0]

    header_map = {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value
    }

    missing_headers = [col for col in REPORT_COLUMNS if col not in header_map]

    if missing_headers:
        raise ValueError(
            f"Missing headers in XLSX template: {', '.join(missing_headers)}"
        )

    first_empty_row = ws.max_row + 1

    for _, row in append_df.iterrows():
        for col_name in REPORT_COLUMNS:
            ws.cell(
                row=first_empty_row,
                column=header_map[col_name],
                value=row[col_name],
            )
        first_empty_row += 1

    output_xlsx = BytesIO()
    wb.save(output_xlsx)
    output_xlsx.seek(0)

    output_apexa_csv = BytesIO()
    apexa_update_df.to_csv(output_apexa_csv, index=False, encoding="utf-8")
    output_apexa_csv.seek(0)

    return (
        output_xlsx,
        output_apexa_csv,
        append_df,
        overdue_result_df,
        pending_updates_df,
        apexa_update_df,
        total_open,
        total_closed,
        pending_to_apply,
    )


st.set_page_config(
    page_title="zOS Patching Report",
    layout="wide",
)

st.title("zOS Patching Report")
st.caption("Developed by Matheus Giroto - Cybersecurity Infrastructure Lead")
st.write("Upload do report XLSX atual + CSV atualizado.")

report_file = st.file_uploader("Current report XLSX", type=["xlsx"])
csv_file = st.file_uploader("Updated CSV", type=["csv"])

if report_file and csv_file:
    if st.button("Generate report"):
        progress = st.progress(0)
        status = st.empty()

        try:
            status.info("Reading files...")
            progress.progress(20)

            (
                output_xlsx,
                output_apexa_csv,
                new_releases,
                overdues,
                pending_updates,
                apexa_update,
                total_open,
                total_closed,
                pending_to_apply,
            ) = generate_report(report_file, csv_file)

            status.info("Calculating results...")
            progress.progress(70)

            new_releases_unique = (
                new_releases["cve_id"].nunique()
                if not new_releases.empty
                else 0
            )

            overdues_unique = (
                overdues["cve_id"].nunique()
                if not overdues.empty
                else 0
            )

            progress.progress(100)
            status.success("Report generated successfully.")

            col1, col2, col3, col4, col5, col6 = st.columns(6)

            col1.metric("New Releases", new_releases_unique)
            col2.metric("Overdues", overdues_unique)
            col3.metric("Pending APEXA Updates", len(pending_updates))
            col4.metric("Pending to Apply", pending_to_apply)
            col5.metric("Total Open", total_open)
            col6.metric("Total Closed", total_closed)

            tab1, tab2, tab3, tab4 = st.tabs(
                [
                    "New Releases",
                    "Overdues",
                    "Pending APEXA Updates",
                    "APEXA Update CSV",
                ]
            )

            with tab1:
                st.subheader("New Releases")
                st.caption("CVEs presentes no CSV e ausentes no report atual.")
                st.dataframe(new_releases, use_container_width=True)

            with tab2:
                st.subheader("Overdues")
                st.caption(
                    "Somente PENDING. SLA: Critical 14 days, High 60 days, Medium 180 days, Low No Target Date."
                )
                st.dataframe(overdues, use_container_width=True)

            with tab3:
                st.subheader("Pending APEXA Updates")
                st.caption(
                    "Itens onde remediation_status ou remediation_comment estão diferentes entre CSV e report atual."
                )
                st.dataframe(pending_updates, use_container_width=True)

            with tab4:
                st.subheader("APEXA Update CSV")
                st.caption(
                    "CSV pronto para atualizar o APEXA com status/comment do XLSX atual."
                )
                st.dataframe(apexa_update, use_container_width=True)

            col_download_1, col_download_2 = st.columns(2)

            with col_download_1:
                st.download_button(
                    label="Download new_report.xlsx",
                    data=output_xlsx,
                    file_name="zOS Monthly Patching Report MM_YYYY.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            with col_download_2:
                st.download_button(
                    label="Download apexa_update.csv",
                    data=output_apexa_csv,
                    file_name="apexa_update.csv",
                    mime="text/csv",
                )

        except Exception as e:
            progress.progress(0)
            status.error(str(e))
