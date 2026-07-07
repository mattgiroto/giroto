# =============================================================================
# zOS Security Automation Suite
# Author / Creator: Matheus Giroto
# Role: Cybersecurity Infrastructure Lead
# Created: 2026
#
# Internal watermark:
# This solution, architecture, workflow and automation logic were designed and
# developed by Matheus Giroto for IBM Z / zOS Vulnerability Management reporting,
# readiness validation and KPI analysis.
# =============================================================================

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime, date, timedelta


APP_NAME = "zOS Security Automation Suite"
APP_AUTHOR = "Matheus Giroto"
APP_ROLE = "Cybersecurity Infrastructure Lead"
APP_WATERMARK = "Designed and Developed by Matheus Giroto"
APP_VERSION = "1.0.0"

CSV_OS_ALLOWED = ["ibm z/os", "ibm z/vm"]

REPORT_COLUMNS = [
    "APAR/TPF/FMID (IBM)",
    "Severity (IBM)",
    "CVSS (IBM)",
    "Due Date (AXP)",
    "Escalation (AXP)",
    "sec_uuid",
    "cve_id",
    "host_name",
    "operating_system",
    "cvss",
    "cve_cache.published",
    "remediation_status",
    "remediation_comment",
    "Team",
]

APEXA_COLUMNS = [
    "cve_id",
    "remediation_status",
    "host_name",
    "device_uuid",
    "msp_tenant_uid",
    "remediation_comment",
]

READINESS_REQUIRED_COLUMNS = [
    "sec_uuid",
    "cve_id",
    "host_name",
    "cvss",
    "cve_cache.published",
    "remediation_status",
    "remediation_comment",
    "Team",
]

VALID_STATUSES = {"PENDING", "REMEDIATED", "FALSE_POSITIVE"}


# =============================================================================
# Shared utilities - Watermark: Matheus Giroto
# =============================================================================

def normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_date(value):
    if pd.isna(value):
        return ""
    return str(value).split("T")[0].strip()


def normalize_severity(value):
    value = normalize_text(value).upper()
    return "MEDIUM" if value == "MODERATE" else value


def normalize_remediation_status(value):
    value = normalize_text(value).upper()
    if value in ["", "DEFFERED", "DEFERRED"]:
        return "PENDING"
    return value


def normalize_remediation_comment(value):
    value = normalize_text(value)
    if value == "":
        return "Pending Status"
    return value


def parse_date(value):
    value = normalize_date(value)
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def calculate_due_date(published, severity):
    published_date = parse_date(published)
    severity = normalize_severity(severity)
    if not published_date:
        return None
    if severity == "CRITICAL":
        return published_date + timedelta(days=14)
    if severity == "HIGH":
        return published_date + timedelta(days=60)
    if severity == "MEDIUM":
        return published_date + timedelta(days=180)
    return None


def calculate_slo_zone(published, severity):
    due_date = calculate_due_date(published, severity)
    if not due_date:
        return "No Due Date"
    today = date.today()
    red_zone_date = due_date + timedelta(days=30)
    if today <= due_date:
        return "Within SLO"
    if today <= red_zone_date:
        return "Danger Zone"
    return "Red Zone"


def validate_columns(df, required_columns, file_name="file"):
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise ValueError(f"Missing columns in {file_name}: {', '.join(missing)}")


def get_team(affected_products):
    text = normalize_text(affected_products).lower()
    has_mvs = (
        "cpe:2.3:o:ibm:z\\/os:" in text
        or "cpe:2.3:o:ibm:z/os:" in text
        or "z\\/os_security" in text
        or "z/os_security" in text
        or "z\\/os_system_automation" in text
        or "z/os_system_automation" in text
        or "z\\/os_storage_software" in text
        or "z/os_storage_software" in text
    )
    has_omegamon = "z\\/os_performance" in text or "z/os_performance" in text
    has_mq = "z\\/os_websphere_mq" in text or "z/os_websphere_mq" in text
    has_cics = "z\\/os_cics" in text or "z/os_cics" in text
    has_db2 = "z\\/os_db2" in text or "z/os_db2" in text
    has_ims = "z\\/os_ims" in text or "z/os_ims" in text
    has_zvm = "cpe:2.3:a:ibm:z\\/vm" in text or "cpe:2.3:a:ibm:z/vm" in text

    if has_db2 and has_ims:
        return "zOS DB2 & IMS"
    if has_mvs and has_omegamon:
        return "zOS MVS & zOS Omegamon"
    if has_mq:
        return "zOS MQ"
    if has_cics:
        return "zOS CICS"
    if has_db2:
        return "zOS DB2"
    if has_ims:
        return "zOS IMS"
    if has_zvm:
        return "zVM"
    if has_mvs:
        return "zOS MVS"
    if has_omegamon:
        return "zOS Omegamon"
    return "UNKNOWN"


# =============================================================================
# Module 1 - zOS Patching Report - Watermark: Matheus Giroto
# =============================================================================

def report_is_overdue(published, severity):
    due_date = calculate_due_date(published, severity)
    if not due_date:
        return False
    return date.today() > due_date


def report_build_output_row(row):
    cve_id = normalize_text(row.get("cve_id"))
    host_name = normalize_text(row.get("host_name"))
    severity = normalize_severity(row.get("severity_name"))
    published = normalize_date(row.get("cve_cache.published"))
    due_date = calculate_due_date(published, severity)
    return {
        "APAR/TPF/FMID (IBM)": "",
        "Severity (IBM)": "",
        "CVSS (IBM)": "",
        "Due Date (AXP)": due_date.strftime("%Y-%m-%d") if due_date else "",
        "Escalation (AXP)": calculate_slo_zone(published, severity),
        "sec_uuid": f"{cve_id}_{host_name}",
        "cve_id": cve_id,
        "host_name": host_name,
        "operating_system": normalize_text(row.get("operating_system")),
        "cvss": severity,
        "cve_cache.published": published,
        "remediation_status": normalize_remediation_status(row.get("remediation_status")),
        "remediation_comment": normalize_remediation_comment(row.get("remediation_comment")),
        "Team": get_team(row.get("affected_products")),
    }


def generate_report(report_file, csv_file):
    report_df = pd.read_excel(report_file, sheet_name=0, dtype=str)
    csv_df = pd.read_csv(csv_file, dtype=str)
    report_df.columns = report_df.columns.str.strip()
    csv_df.columns = csv_df.columns.str.strip()

    validate_columns(report_df, ["cve_id", "remediation_status", "remediation_comment"], "current report XLSX")
    validate_columns(csv_df, [
        "cve_id", "host_name", "operating_system", "severity_name",
        "cve_cache.published", "remediation_status", "remediation_comment", "affected_products"
    ], "CSV")

    csv_df["operating_system_norm"] = csv_df["operating_system"].fillna("").str.strip().str.lower()
    csv_df = csv_df[csv_df["operating_system_norm"].isin(CSV_OS_ALLOWED)].copy()
    report_df["cve_id_norm"] = report_df["cve_id"].fillna("").astype(str).str.strip()
    csv_df["cve_id_norm"] = csv_df["cve_id"].fillna("").astype(str).str.strip()
    csv_df["remediation_status_norm"] = csv_df["remediation_status"].apply(normalize_remediation_status)
    csv_df["remediation_comment_norm"] = csv_df["remediation_comment"].apply(normalize_remediation_comment)
    report_df["remediation_status_norm"] = report_df["remediation_status"].apply(normalize_remediation_status)
    report_df["remediation_comment_norm"] = report_df["remediation_comment"].apply(normalize_remediation_comment)

    existing_cves = set(report_df["cve_id_norm"].dropna())
    new_releases_df = csv_df[(~csv_df["cve_id_norm"].isin(existing_cves)) & (csv_df["cve_id_norm"] != "")].copy()
    report_lookup = report_df.set_index("cve_id_norm")

    pending_rows = []
    apexa_rows = []
    for _, row in csv_df.iterrows():
        cve_id = row["cve_id_norm"]
        if cve_id in report_lookup.index:
            report_row = report_lookup.loc[cve_id]
            if isinstance(report_row, pd.DataFrame):
                report_row = report_row.iloc[0]
            csv_status = normalize_remediation_status(row.get("remediation_status"))
            csv_comment = normalize_remediation_comment(row.get("remediation_comment"))
            report_status = normalize_remediation_status(report_row.get("remediation_status"))
            report_comment = normalize_remediation_comment(report_row.get("remediation_comment"))
            if csv_status != report_status or csv_comment != report_comment:
                pending_rows.append({
                    "cve_id": cve_id,
                    "host_name": normalize_text(row.get("host_name")),
                    "operating_system": normalize_text(row.get("operating_system")),
                    "csv_remediation_status": csv_status,
                    "report_remediation_status": report_status,
                    "csv_remediation_comment": csv_comment,
                    "report_remediation_comment": report_comment,
                })
                apexa_rows.append({
                    "cve_id": cve_id,
                    "remediation_status": report_status,
                    "host_name": normalize_text(row.get("host_name")),
                    "device_uuid": "",
                    "msp_tenant_uid": "",
                    "remediation_comment": report_comment,
                })

    pending_updates_df = pd.DataFrame(pending_rows)
    apexa_update_df = pd.DataFrame(apexa_rows, columns=APEXA_COLUMNS)

    overdue_df = csv_df[(csv_df["remediation_status_norm"] == "PENDING") & csv_df.apply(
        lambda row: report_is_overdue(row.get("cve_cache.published"), row.get("severity_name")), axis=1
    )].copy()

    overdue_rows = []
    for _, row in overdue_df.iterrows():
        severity = normalize_severity(row.get("severity_name"))
        published = normalize_date(row.get("cve_cache.published"))
        due_date = calculate_due_date(published, severity)
        overdue_rows.append({
            "cve_id": normalize_text(row.get("cve_id")),
            "host_name": normalize_text(row.get("host_name")),
            "operating_system": normalize_text(row.get("operating_system")),
            "severity": severity,
            "published": published,
            "due_date": due_date.strftime("%Y-%m-%d") if due_date else "",
            "slo_zone": calculate_slo_zone(published, severity),
            "days_overdue": (date.today() - due_date).days if due_date else "",
            "remediation_status": normalize_remediation_status(row.get("remediation_status")),
            "remediation_comment": normalize_remediation_comment(row.get("remediation_comment")),
            "Team": get_team(row.get("affected_products")),
        })

    overdue_result_df = pd.DataFrame(overdue_rows)
    append_rows = [report_build_output_row(row) for _, row in new_releases_df.iterrows()]
    append_df = pd.DataFrame(append_rows, columns=REPORT_COLUMNS)

    total_open = csv_df[csv_df["remediation_status_norm"] == "PENDING"]["cve_id_norm"].nunique()
    total_closed = csv_df[csv_df["remediation_status_norm"].isin(["REMEDIATED", "FALSE_POSITIVE"])]["cve_id_norm"].nunique()
    pending_to_apply = total_open

    report_file.seek(0)
    wb = load_workbook(report_file)
    ws = wb.worksheets[0]
    header_map = {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value}
    missing_headers = [col for col in REPORT_COLUMNS if col not in header_map]
    if missing_headers:
        raise ValueError(f"Missing headers in XLSX template: {', '.join(missing_headers)}")

    first_empty_row = ws.max_row + 1
    for _, row in append_df.iterrows():
        for col_name in REPORT_COLUMNS:
            ws.cell(row=first_empty_row, column=header_map[col_name], value=row[col_name])
        first_empty_row += 1

    # Workbook metadata watermark
    wb.properties.creator = APP_AUTHOR
    wb.properties.lastModifiedBy = APP_AUTHOR
    wb.properties.title = "zOS Monthly Patching Report"
    wb.properties.subject = APP_WATERMARK
    wb.properties.description = f"Generated by {APP_NAME} - {APP_WATERMARK}"

    output_xlsx = BytesIO()
    wb.save(output_xlsx)
    output_xlsx.seek(0)

    output_apexa_csv = BytesIO()
    apexa_update_df.to_csv(output_apexa_csv, index=False, encoding="utf-8")
    output_apexa_csv.seek(0)

    return output_xlsx, output_apexa_csv, append_df, overdue_result_df, pending_updates_df, apexa_update_df, total_open, total_closed, pending_to_apply


def render_patching_report():
    st.title("zOS Patching Report")
    st.caption(f"{APP_WATERMARK} - {APP_ROLE}")
    st.write("Upload do report XLSX atual + CSV atualizado.")

    report_file = st.file_uploader("Current report XLSX", type=["xlsx"], key="report_xlsx")
    csv_file = st.file_uploader("Updated CSV", type=["csv"], key="report_csv")

    if report_file and csv_file and st.button("Generate report", type="primary"):
        progress = st.progress(0)
        status = st.empty()
        try:
            status.info("Reading files...")
            progress.progress(20)
            output_xlsx, output_apexa_csv, new_releases, overdues, pending_updates, apexa_update, total_open, total_closed, pending_to_apply = generate_report(report_file, csv_file)
            status.info("Calculating results...")
            progress.progress(70)
            new_releases_unique = new_releases["cve_id"].nunique() if not new_releases.empty else 0
            overdues_unique = overdues["cve_id"].nunique() if not overdues.empty else 0
            progress.progress(100)
            status.success("Report generated successfully.")

            col1, col2, col3, col4, col5, col6 = st.columns(6)
            col1.metric("New Releases", new_releases_unique)
            col2.metric("Overdues", overdues_unique)
            col3.metric("Pending APEXA Updates", len(pending_updates))
            col4.metric("Pending to Apply", pending_to_apply)
            col5.metric("Total Open", total_open)
            col6.metric("Total Closed", total_closed)

            tab1, tab2, tab3, tab4 = st.tabs(["New Releases", "Overdues", "Pending APEXA Updates", "APEXA Update CSV"])
            with tab1:
                st.subheader("New Releases")
                st.dataframe(new_releases, use_container_width=True)
            with tab2:
                st.subheader("Overdues")
                st.caption("Critical 14 days, High 60 days, Medium 180 days, Low/Info No Due Date. Danger Zone after SLO; Red Zone after 30 grace days.")
                st.dataframe(overdues, use_container_width=True)
            with tab3:
                st.subheader("Pending APEXA Updates")
                st.dataframe(pending_updates, use_container_width=True)
            with tab4:
                st.subheader("APEXA Update CSV")
                st.dataframe(apexa_update, use_container_width=True)

            col_download_1, col_download_2 = st.columns(2)
            with col_download_1:
                st.download_button("Download new_report.xlsx", data=output_xlsx, file_name="zOS Monthly Patching Report MM_YYYY.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            with col_download_2:
                st.download_button("Download apexa_update.csv", data=output_apexa_csv, file_name="apexa_update.csv", mime="text/csv")
        except Exception as e:
            progress.progress(0)
            status.error(str(e))


# =============================================================================
# Module 2 - XLSX Readiness Analyzer - Watermark: Matheus Giroto
# =============================================================================

def readiness_normalize_status(value):
    return normalize_text(value).upper()


def readiness_is_overdue(row):
    status = readiness_normalize_status(row.get("remediation_status"))
    if status != "PENDING":
        return False
    due_date = calculate_due_date(row.get("cve_cache.published"), row.get("cvss"))
    if not due_date:
        return False
    return date.today() > due_date


def add_issue(issue_details, check, excel_row, cve_id, host_name, current_value, expected_value, message):
    issue_details.append({
        "check": check,
        "excel_row": excel_row,
        "cve_id": cve_id,
        "host_name": host_name,
        "current_value": current_value,
        "expected_value": expected_value,
        "message": message,
    })


def analyze_xlsx(uploaded_file):
    df = pd.read_excel(uploaded_file, dtype=str)
    df.columns = df.columns.str.strip()
    total_rows = len(df)
    issues, warnings, issue_details, warning_details = [], [], [], []
    df["excel_row"] = df.index + 2

    missing_columns = [col for col in READINESS_REQUIRED_COLUMNS if col not in df.columns]
    if missing_columns:
        issues.append({"check": "Missing required columns", "severity": "BLOCKER", "count": len(missing_columns), "rows": "N/A", "details": ", ".join(missing_columns)})
        for col in missing_columns:
            add_issue(issue_details, "Missing required columns", "N/A", "", "", "Missing", col, f"Required column is missing: {col}")
        return df, issues, warnings, issue_details, warning_details, 0

    df["sec_uuid_norm"] = df["sec_uuid"].apply(normalize_text)
    df["cve_norm"] = df["cve_id"].apply(normalize_text)
    df["host_norm"] = df["host_name"].apply(normalize_text)
    df["remediation_status_norm"] = df["remediation_status"].apply(readiness_normalize_status)
    df["remediation_comment_norm"] = df["remediation_comment"].apply(normalize_text)
    df["cvss_norm"] = df["cvss"].apply(lambda x: normalize_text(x).upper())
    df["team_norm"] = df["Team"].apply(normalize_text)
    expected_sec_uuid = df["cve_norm"] + "_" + df["host_norm"]

    checks = [
        ("Blank sec_uuid", df[df["sec_uuid_norm"] == ""], "sec_uuid", "", "cve_id_host_name", "Rows with blank sec_uuid."),
        ("Invalid sec_uuid", df[(df["sec_uuid_norm"] != "") & (df["sec_uuid_norm"] != expected_sec_uuid)], "sec_uuid", "dynamic_expected_sec_uuid", "sec_uuid must follow this format: cve_id_host_name.", "Invalid sec_uuid format."),
        ("Duplicated sec_uuid", df[(df["sec_uuid_norm"] != "") & df["sec_uuid_norm"].duplicated(keep=False)], "sec_uuid", "Unique sec_uuid", "Duplicated sec_uuid values detected.", "Duplicated sec_uuid detected."),
        ("Blank remediation_status", df[df["remediation_status_norm"] == ""], "remediation_status", "", "PENDING, REMEDIATED, or FALSE_POSITIVE", "Rows with blank remediation_status."),
        ("Blank remediation_comment", df[df["remediation_comment_norm"] == ""], "remediation_comment", "", "Non-blank remediation_comment", "Rows with blank remediation_comment."),
        ("Invalid status", df[~df["remediation_status_norm"].isin(VALID_STATUSES)], "remediation_status", "PENDING, REMEDIATED, or FALSE_POSITIVE", "Allowed values: PENDING, REMEDIATED, FALSE_POSITIVE.", "Invalid remediation_status."),
        ("MODERATE not converted to MEDIUM", df[df["cvss_norm"] == "MODERATE"], "cvss", "MEDIUM", "Rows still using MODERATE instead of MEDIUM.", "Severity must be MEDIUM instead of MODERATE."),
        ("Date contains time in cve_cache.published", df[df["cve_cache.published"].fillna("").astype(str).str.contains("T", regex=False)], "cve_cache.published", "YYYY-MM-DD", "Date should be YYYY-MM-DD only.", "Date contains time component."),
        ("Team UNKNOWN or blank", df[(df["team_norm"] == "") | (df["team_norm"].str.upper() == "UNKNOWN")], "Team", "Known team name", "Rows with Team UNKNOWN or blank.", "Team is UNKNOWN or blank."),
    ]

    for check_name, bad_df, column_name, expected_value, details, message in checks:
        if bad_df.empty:
            continue
        row_numbers = bad_df["excel_row"].astype(str).tolist()
        issues.append({"check": check_name, "severity": "BLOCKER", "count": len(bad_df), "rows": ", ".join(row_numbers[:50]) + ("..." if len(row_numbers) > 50 else ""), "details": details})
        for _, row in bad_df.iterrows():
            if check_name == "Invalid sec_uuid":
                current = row.get("sec_uuid", "")
                expected = f"{row.get('cve_norm', '')}_{row.get('host_norm', '')}"
            else:
                current = row.get(column_name, "")
                expected = expected_value
            add_issue(issue_details, check_name, row.get("excel_row"), row.get("cve_id", ""), row.get("host_name", ""), current, expected, message)

    df["is_overdue"] = df.apply(readiness_is_overdue, axis=1)
    overdue_pending = df[df["is_overdue"] == True]
    if not overdue_pending.empty:
        row_numbers = overdue_pending["excel_row"].astype(str).tolist()
        warnings.append({"check": "Pending overdue patches", "severity": "WARNING", "count": len(overdue_pending), "rows": ", ".join(row_numbers[:50]) + ("..." if len(row_numbers) > 50 else ""), "details": f"File can be submitted, but there are {len(overdue_pending)} pending overdue patches."})
        for _, row in overdue_pending.iterrows():
            add_issue(warning_details, "Pending overdue patches", row.get("excel_row"), row.get("cve_id", ""), row.get("host_name", ""), row.get("remediation_status", ""), "Not a blocker", "File can be submitted, but this PENDING item is overdue.")

    total_blockers = sum(item["count"] for item in issues)
    readiness = 0 if total_rows == 0 else max(0, round(100 - ((total_blockers / total_rows) * 100), 2))
    return df, issues, warnings, issue_details, warning_details, readiness


def render_readiness_analyzer():
    st.title("XLSX Readiness Analyzer")
    st.caption(f"{APP_WATERMARK} | Report Submission Quality Check")
    st.write("Upload the XLSX report to validate whether it is ready to be submitted.")

    uploaded_file = st.file_uploader("Upload XLSX", type=["xlsx"], key="readiness_xlsx")
    if uploaded_file and st.button("Analyze File", type="primary"):
        try:
            df, issues, warnings, issue_details, warning_details, readiness = analyze_xlsx(uploaded_file)
            st.subheader("Readiness Status")
            st.progress(int(readiness))
            if readiness == 100:
                st.warning(f"File is ready to be submitted, but there are {warnings[0]['count']} pending overdue patches.") if warnings else st.success("File is 100% ready to be submitted.")
            else:
                st.error(f"File is {readiness}% ready to be submitted.")

            col1, col2, col3 = st.columns(3)
            col1.metric("Readiness", f"{readiness}%")
            col2.metric("Blocking Issues", sum(item["count"] for item in issues))
            col3.metric("Warnings", sum(item["count"] for item in warnings))

            tab1, tab2, tab3, tab4, tab5 = st.tabs(["Blocking Summary", "Issue Details", "Warnings", "sec_uuid Details", "Raw Data"])
            with tab1:
                st.subheader("Blocking Summary")
                st.dataframe(pd.DataFrame(issues), use_container_width=True) if issues else st.success("No blocking issues found.")
            with tab2:
                st.subheader("Issue Details by Excel Row")
                if issue_details:
                    details_df = pd.DataFrame(issue_details)
                    st.dataframe(details_df, use_container_width=True)
                    st.download_button("Download Issue Details CSV", data=details_df.to_csv(index=False).encode("utf-8"), file_name="xlsx_readiness_issue_details.csv", mime="text/csv")
                else:
                    st.success("No issue details found.")
            with tab3:
                st.subheader("Warnings")
                if warnings:
                    st.dataframe(pd.DataFrame(warnings), use_container_width=True)
                    if warning_details:
                        st.subheader("Warning Details by Excel Row")
                        st.dataframe(pd.DataFrame(warning_details), use_container_width=True)
                else:
                    st.success("No warnings found.")
            with tab4:
                st.subheader("sec_uuid Validation Details")
                if "sec_uuid_norm" in df.columns:
                    sec_uuid_details = df.copy()
                    sec_uuid_details["expected_sec_uuid"] = sec_uuid_details["cve_norm"] + "_" + sec_uuid_details["host_norm"]
                    sec_uuid_details["sec_uuid_valid"] = sec_uuid_details["sec_uuid_norm"] == sec_uuid_details["expected_sec_uuid"]
                    st.dataframe(sec_uuid_details[["excel_row", "sec_uuid", "expected_sec_uuid", "sec_uuid_valid", "cve_id", "host_name"]], use_container_width=True)
                else:
                    st.info("sec_uuid validation could not run due to missing required columns.")
            with tab5:
                st.subheader("Raw Data")
                st.dataframe(df, use_container_width=True)
        except Exception as e:
            st.error(str(e))


# =============================================================================
# Module 3 - zOS Patching Analyzer / KPI Dashboard - Watermark: Matheus Giroto
# =============================================================================

def analyzer_is_overdue(row):
    if row["remediation_status_norm"] != "PENDING":
        return False
    due_date = calculate_due_date(row.get("cve_cache.published"), row.get("cvss"))
    if not due_date:
        return False
    return date.today() > due_date


def analyzer_prepare_data(xlsx_file):
    df = pd.read_excel(xlsx_file, sheet_name=0, dtype=str)
    df.columns = df.columns.str.strip()
    validate_columns(df, ["cve_id", "host_name", "operating_system", "cvss", "cve_cache.published", "remediation_status", "Team"], "XLSX")
    df["operating_system_norm"] = df["operating_system"].fillna("").str.strip().str.lower()
    df = df[df["operating_system_norm"].isin(CSV_OS_ALLOWED)].copy()
    df["cve_id_norm"] = df["cve_id"].fillna("").astype(str).str.strip()
    df = df[df["cve_id_norm"] != ""].copy()
    df["severity_norm"] = df["cvss"].apply(normalize_severity)
    df["remediation_status_norm"] = df["remediation_status"].apply(normalize_remediation_status)
    df["published_date"] = df["cve_cache.published"].apply(parse_date)
    if "remediation_status_updated_at" in df.columns:
        df["status_updated_date"] = df["remediation_status_updated_at"].apply(parse_date)
    else:
        df["status_updated_date"] = None
    df["due_date"] = df.apply(lambda r: calculate_due_date(r.get("cve_cache.published"), r.get("cvss")), axis=1)
    df["slo_zone"] = df.apply(lambda r: calculate_slo_zone(r.get("cve_cache.published"), r.get("cvss")), axis=1)
    df["is_overdue"] = df.apply(analyzer_is_overdue, axis=1)
    return df


def analyzer_build_kpis(df):
    installed_patches = df[df["remediation_status_norm"] == "REMEDIATED"]["cve_id_norm"].nunique()
    pending_patches = df[df["remediation_status_norm"] == "PENDING"]["cve_id_norm"].nunique()
    critical_overdues = df[(df["severity_norm"] == "CRITICAL") & (df["is_overdue"] )]["cve_id_norm"].nunique()
    high_overdues = df[(df["severity_norm"] == "HIGH") & (df["is_overdue"] )]["cve_id_norm"].nunique()
    medium_overdues = df[(df["severity_norm"] == "MEDIUM") & (df["is_overdue"] )]["cve_id_norm"].nunique()
    last_30_days = date.today() - timedelta(days=30)
    installed_last_month = df[(df["remediation_status_norm"] == "REMEDIATED") & (df["status_updated_date"].notna()) & (df["status_updated_date"] >= last_30_days)]["cve_id_norm"].nunique()
    danger_zone = df[(df["remediation_status_norm"] == "PENDING") & (df["slo_zone"] == "Danger Zone")]["cve_id_norm"].nunique()
    red_zone = df[(df["remediation_status_norm"] == "PENDING") & (df["slo_zone"] == "Red Zone")]["cve_id_norm"].nunique()
    return {
        "installed_patches": installed_patches,
        "pending_patches": pending_patches,
        "critical_overdues": critical_overdues,
        "high_overdues": high_overdues,
        "medium_overdues": medium_overdues,
        "installed_last_month": installed_last_month,
        "danger_zone": danger_zone,
        "red_zone": red_zone,
    }


def analyzer_build_pending_by_team(df):
    pending_df = df[df["remediation_status_norm"] == "PENDING"].copy()
    return pending_df.groupby("Team")["cve_id_norm"].nunique().reset_index().rename(columns={"cve_id_norm": "pending_cves"}).sort_values("pending_cves", ascending=False)


def render_patching_analyzer():
    st.title("zOS KPI Dashboard / Patching Analyzer")
    st.caption(f"{APP_WATERMARK} | IBM Z Vulnerability Management Automation")
    st.write("Upload do report XLSX para visualizar KPIs de patching.")

    xlsx_file = st.file_uploader("Upload XLSX", type=["xlsx"], key="analyzer_xlsx")
    if xlsx_file and st.button("Generate KPIs", type="primary"):
        progress = st.progress(0)
        status = st.empty()
        try:
            status.info("Reading XLSX...")
            progress.progress(25)
            df = analyzer_prepare_data(xlsx_file)
            status.info("Calculating KPIs...")
            progress.progress(70)
            kpis = analyzer_build_kpis(df)
            pending_by_team = analyzer_build_pending_by_team(df)
            progress.progress(100)
            status.success("KPIs generated successfully.")

            col1, col2, col3 = st.columns(3)
            col1.metric("Patches Installed", kpis["installed_patches"])
            col2.metric("Pending Patches", kpis["pending_patches"])
            col3.metric("Installed Last 30 Days", kpis["installed_last_month"])
            col4, col5, col6 = st.columns(3)
            col4.metric("Critical Overdues", kpis["critical_overdues"])
            col5.metric("High Overdues", kpis["high_overdues"])
            col6.metric("Medium Overdues", kpis["medium_overdues"])
            col7, col8 = st.columns(2)
            col7.metric("Danger Zone", kpis["danger_zone"])
            col8.metric("Red Zone", kpis["red_zone"])

            tab1, tab2, tab3 = st.tabs(["Pending by Team", "Overdue Details", "Raw XLSX Data"])
            with tab1:
                st.subheader("Pending Patches by Team")
                st.dataframe(pending_by_team, use_container_width=True)
                if not pending_by_team.empty:
                    st.bar_chart(pending_by_team.set_index("Team")["pending_cves"])
            with tab2:
                st.subheader("Overdue Details")
                overdue_details = df[df["is_overdue"]].copy()
                columns = ["cve_id", "host_name", "operating_system", "severity_norm", "cve_cache.published", "due_date", "slo_zone", "remediation_status_norm", "Team"]
                overdue_details = overdue_details[columns].rename(columns={"severity_norm": "severity", "remediation_status_norm": "remediation_status"})
                st.dataframe(overdue_details, use_container_width=True)
            with tab3:
                st.subheader("Filtered XLSX Data")
                st.dataframe(df, use_container_width=True)
        except Exception as e:
            progress.progress(0)
            status.error(str(e))


# =============================================================================
# Main application - Watermark: Matheus Giroto
# =============================================================================

def render_footer():
    st.divider()
    st.caption(f"{APP_NAME} v{APP_VERSION} | {APP_WATERMARK} | {APP_ROLE}")


def main():
    st.set_page_config(page_title=APP_NAME, layout="wide")

    with st.sidebar:
        st.title("IBM Z Security Suite")
        st.caption(APP_WATERMARK)
        st.caption(f"Version {APP_VERSION}")
        selected_app = st.radio(
            "Choose the module",
            [
                "zOS Patching Report",
                "XLSX Readiness Analyzer",
                "zOS Patching Analyzer / KPIs",
            ],
        )
        st.divider()
        st.markdown("**Author**")
        st.write(APP_AUTHOR)
        st.markdown("**Role**")
        st.write(APP_ROLE)
        st.markdown("**SLO Policy**")
        st.write("Critical: 14 days")
        st.write("High: 60 days")
        st.write("Medium: 180 days")
        st.write("Low/Info: No Due Date")
        st.write("30 days grace after original SLO")

    if selected_app == "zOS Patching Report":
        render_patching_report()
    elif selected_app == "XLSX Readiness Analyzer":
        render_readiness_analyzer()
    else:
        render_patching_analyzer()

    render_footer()


if __name__ == "__main__":
    main()
